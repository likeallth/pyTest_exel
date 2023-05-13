import openpyxl as op

#path는 대상이 되는 엑셀파일의 절대 경로
def corretion(path : str):
    #openpyxl 워크북 정의
    wb = op.load_workbook(path)
    #워크북의 활성화된 시트를 객체로 정의
    ws = wb.active

    #해당 시트의 마지막 열, 마지막 행 
    column_max = ws.max_column
    row_max = ws.max_row

    #열마다 행을 for loop문 진행
    for col_num in range(1, column_max+1):
        for row_num in range(1, row_max+1):
            
            #tempstr : cell값이 문자열이 아닌 경우를 감안하여 str로 바꿔줌
            tempstr = str(ws.cell(row = row_num, column = col_num).value)
            #문자열 함수 replace 사용(,를 빈칸으로)
            data = tempstr.replace(",","")
            #빈 셀의 경우 None이라는 문자열 타입이므로 제외하고 데이터 입력
            if data != "None":
                ws.cell(row = row_num, column = col_num).value = data 
    
    #저장
    wb.save("correction_result.xlsx")

if __name__ == "__main__":
    path = r"엑셀 파일 절대 경로"

    #함수 실행
    corretion(path)