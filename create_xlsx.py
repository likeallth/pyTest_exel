from openpyxl import Workbook, load_workbook

class AutoExelIo():
    def __init__(self, FileName):
        #self.ExelFileName = 'My_document'
        #self.ExelSheetName = 'Summary'
        self.wb = Workbook(FileName)
        
    def Create(self, FileName):#, SheetName):
        #wb = Workbook()
        #ws = self.wb.active
        self.ExelSheetName = FileName
        #self.ExelSheetName = SheetName
        #ws.title = self.ExelSheetName #qt ui xlxs 생성 파일 이름 칸과 연결
        self.wb.save(self.ExelFileName) #
        
    def AddSheet(self, SheetName):
        #ws = self.wb.active()
        self.wb.create_sheet(SheetName)
        
    def DelSheet(self, SheetName):
       del self.wb[SheetName]
    
    def FindandInsert(self, srcName, dstName, rowNum, colNum):
        
        
    def CloseExel(self):
        self.wb.close()